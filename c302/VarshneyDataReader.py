from c302.ConnectomeReader import ConnectionInfo
from c302.ConnectomeReader import analyse_connections
from openpyxl import load_workbook

import os
from c302 import print_

spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"
spreadsheet_name = "NeuronConnect.xlsx"  # has old name...
spreadsheet_name = "NeuronConnectFormatted.xlsx"

READER_DESCRIPTION = (
    """Data extracted from %s for neuronal connectivity""" % spreadsheet_name
)

NMJ_ENDPOINT = "NMJ"


def read_data(include_nonconnected_cells=False, neuron_connect=True):
    if neuron_connect:
        conns = []
        cells = []
        filename = "%s%s" % (spreadsheet_location, spreadsheet_name)
        wb = load_workbook(filename)
        sheet = wb.worksheets[0]
        print_("Opened the Excel file: " + filename)

        for row in sheet.iter_rows(
            min_row=2, values_only=True
        ):  # Assuming data starts from the second row
            pre = str(row[0])
            post = str(row[1])

            if not post == NMJ_ENDPOINT:
                syntype = str(row[2])
                num = int(row[3])
                synclass = "Generic_GJ" if "EJ" in syntype else "Generic_CS"

                conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
                if pre not in cells:
                    cells.append(pre)
                if post not in cells:
                    cells.append(post)

        return cells, conns


def read_muscle_data():
    conns = []
    neurons = []
    muscles = []

    filename = "%s%s" % (spreadsheet_location, spreadsheet_name)
    wb = load_workbook(filename)
    sheet = wb.worksheets[0]

    print_("Opened Excel file: " + filename)

    for row in sheet.iter_rows(
        min_row=2, values_only=True
    ):  # Assuming data starts from the second row
        pre = str(row[0])
        post = str(row[1])

        if post == NMJ_ENDPOINT:
            syntype = str(row[2])
            num = int(row[3])
            synclass = "Generic_GJ" if "EJ" in syntype else "Generic_CS"

            conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
            if pre not in neurons:
                neurons.append(pre)
            if post not in muscles:
                muscles.append(post)

    return neurons, muscles, conns


def main():
    cells, neuron_conns = read_data(include_nonconnected_cells=True)
    neurons2muscles, muscles, muscle_conns = read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)


if __name__ == "__main__":
    main()
