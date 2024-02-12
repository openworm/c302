from c302.NeuroMLUtilities import ConnectionInfo, analyse_connections
from openpyxl import load_workbook
import os
from c302 import print_

class ExcelDataReader:
    def __init__(self, file_prefix, num_files=8):
        self.file_prefix = file_prefix
        self.num_files = num_files
        self.spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"

    def read_data(self, include_nonconnected_cells=False, neuron_connect=False):
        cells = []
        conns = []

        for i in range(1, self.num_files + 1):
            filename = f"{self.spreadsheet_location}{self.file_prefix}_{i}.xlsx"
            wb = load_workbook(filename)
            sheet = wb.worksheets[0]
            print_("Opened the Excel file: " + filename)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                pre = str(row[0])
                post = str(row[1])
                syntype = str(row[2])
                num = int(row[3])
                synclass = 'Generic_GJ' if 'electrical' in syntype else 'Chemical_Synapse'

                conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
                if pre not in cells:
                    cells.append(pre)
                if post not in cells:
                    cells.append(post)

        if include_nonconnected_cells:
            known_nonconnected_cells = ['CANL', 'CANR', 'VC6']
            for c in known_nonconnected_cells:
                cells.append(c)

        return cells, conns

    def read_muscle_data(self):
        neurons = []
        muscles = []
        conns = []

        for i in range(1, self.num_files + 1):
            filename = f"{self.spreadsheet_location}{self.file_prefix}_{i}.xlsx"
            wb = load_workbook(filename)
            sheet = wb.worksheets[0]
            print_("Opened Excel file: " + filename)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                pre = str(row[0])
                post = str(row[1])
                syntype = str(row[2])
                num = int(row[3])
                synclass = 'Generic_GJ' if 'electrical' in syntype else 'Chemical_Synapse'

                conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
                if pre not in neurons:
                    neurons.append(pre)
                if post not in muscles:
                    muscles.append(post)

        return neurons, muscles, conns

    def main(self):
        cells, neuron_conns = self.read_data(include_nonconnected_cells=True)
        neurons2muscles, muscles, muscle_conns = self.read_muscle_data()
        analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)


if __name__ == '__main__':
    excel_reader = ExcelDataReader(file_prefix='witvliet_2020', num_files=8)
    excel_reader.main()
