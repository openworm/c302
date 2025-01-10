from c302.ConnectomeReader import ConnectionInfo
from c302.ConnectomeReader import analyse_connections
from c302.ConnectomeReader import convert_to_preferred_muscle_name
from c302.ConnectomeReader import is_neuron

from openpyxl import load_workbook
import os
from c302 import print_


spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"


class WitvlietDataReader:
    def __init__(self, spreadsheet):
        self.filename = "%s%s" % (spreadsheet_location, spreadsheet)

    def read_data(self, include_nonconnected_cells=True):
        if not include_nonconnected_cells:
            raise Exception("Option include_nonconnected_cells=False not supported")

        conns = []
        cells = []

        wb = load_workbook(self.filename)
        sheet = wb.worksheets[0]
        print_("Opened the Excel file: " + self.filename)

        for row in sheet.iter_rows(
            min_row=2, values_only=True
        ):  # Assuming data starts from the second row
            # print(row)
            pre = str(row[0])
            post = str(row[1])

            if not is_neuron(pre) or not is_neuron(post):
                continue  # pre or post is not a neuron

            syntype = str(row[2])
            num = int(row[3])
            synclass = "Generic_GJ" if "electrical" in syntype else "Generic_CS"

            conns.append(ConnectionInfo(pre, post, num, syntype, synclass))

            if pre not in cells:
                cells.append(pre)
            if post not in cells:
                cells.append(post)

        return cells, conns

    def read_muscle_data(self):
        conns = []
        neurons = []
        muscles = []

        wb = load_workbook(self.filename)
        sheet = wb.worksheets[0]

        print_("Opened Excel file: " + self.filename)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            pre = str(row[0])
            post = str(row[1])
            syntype = str(row[2])
            num = int(row[3])
            synclass = "Generic_GJ" if "EJ" in syntype else "Generic_CS"

            if post.startswith("BWM-"):
                post = convert_to_preferred_muscle_name(post)
            else:
                continue

            conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
            if pre not in neurons:
                neurons.append(pre)
            if post not in muscles:
                muscles.append(post)

        return neurons, muscles, conns


def main1():
    wdr = WitvlietDataReader("witvliet_2020_7.xlsx")

    cells, neuron_conns = wdr.read_data()
    neurons2muscles, muscles, muscle_conns = wdr.read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)


def main2():
    wdr = WitvlietDataReader("witvliet_2020_8.xlsx")

    cells, neuron_conns = wdr.read_data()
    neurons2muscles, muscles, muscle_conns = wdr.read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)


if __name__ == "__main__":
    main1()
    main2()
