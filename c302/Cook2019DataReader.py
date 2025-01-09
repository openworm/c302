# -*- coding: utf-8 -*-

############################################################

#    A simple script to read the values in herm_full_edgelist.csv.

#    This is on of a number of interchangeable "Readers" which can
#    be used to get connection data for c302

############################################################


from c302.ConnectomeReader import ConnectionInfo
from c302.ConnectomeReader import analyse_connections
from c302.ConnectomeReader import is_body_wall_muscle
from c302.ConnectomeReader import remove_leading_index_zero

from openpyxl import load_workbook

import os
import numpy as np

from c302 import print_

spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"
filename = "%sSI 5 Connectome adjacency matrices.xlsx" % spreadsheet_location

HERM_CHEM = "hermaphrodite chemical"
HERM_GAP_SYMM = "herm gap jn symmetric"

pre_range = {HERM_CHEM: range(4, 304), HERM_GAP_SYMM: range(4, 472)}

post_range = {HERM_CHEM: range(4, 457), HERM_GAP_SYMM: range(4, 472)}


def get_synclass(cell, syntype):
    # TODO: fix this dirty hack
    if syntype == "GapJunction":
        return "Generic_GJ"
    else:
        if cell.startswith("DD") or cell.startswith("VD"):
            return "GABA"
        return "Acetylcholine"


class Cook2019DataReader:
    verbose = False

    def __init__(self):
        wb = load_workbook(filename)
        print_("Opened the Excel file: " + filename)

        self.pre_cells = {}
        self.post_cells = {}
        self.conn_nums = {}

        for conn_type in [HERM_CHEM, HERM_GAP_SYMM]:
            sheet = wb.get_sheet_by_name(conn_type)
            print("Looking at sheet: %s" % conn_type)

            self.pre_cells[conn_type] = []
            self.post_cells[conn_type] = []

            for i in pre_range[conn_type]:
                self.pre_cells[conn_type].append(sheet["C%i" % i].value)

            if self.verbose:
                print(
                    " - Pre cells for %s (%i):\n%s"
                    % (
                        conn_type,
                        len(self.pre_cells[conn_type]),
                        self.pre_cells[conn_type],
                    )
                )

            for i in post_range[conn_type]:
                self.post_cells[conn_type].append(sheet.cell(row=3, column=i).value)

            if self.verbose:
                print(
                    " - Post cells for %s (%i):\n%s"
                    % (
                        conn_type,
                        len(self.post_cells[conn_type]),
                        self.post_cells[conn_type],
                    )
                )

            self.conn_nums[conn_type] = np.zeros(
                [len(self.pre_cells[conn_type]), len(self.post_cells[conn_type])],
                dtype=int,
            )

            for i in range(len(self.pre_cells[conn_type])):
                for j in range(len(self.post_cells[conn_type])):
                    row = 4 + i
                    col = 4 + j
                    val = sheet.cell(row=row, column=col).value
                    # print('Cell (%i,%i) [row %i, col %i] = %s'%(i,j,row, col, val))
                    if val is not None:
                        self.conn_nums[conn_type][i, j] = int(val)

            if self.verbose:
                print(
                    " - Conns for %s (%s):\n%s"
                    % (
                        conn_type,
                        self.conn_nums[conn_type].shape,
                        self.conn_nums[conn_type],
                    )
                )

    def read_data(self, include_nonconnected_cells=False):
        """
        Args:
            include_nonconnected_cells (bool): Also append neurons without known connections to other neurons to the 'cells' list. True if they should get appended, False otherwise.
        Returns:
            cells (:obj:`list` of :obj:`str`): List of neurons
            conns (:obj:`list` of :obj:`ConnectionInfo`): List of connections from neuron to neuron
        """

        if not include_nonconnected_cells:
            raise Exception("Option include_nonconnected_cells=False not supported")

        conns = []
        cells = []
        for conn_type in [HERM_CHEM, HERM_GAP_SYMM]:
            for pre_index in range(len(self.pre_cells[conn_type])):
                for post_index in range(len(self.post_cells[conn_type])):
                    pre = remove_leading_index_zero(
                        self.pre_cells[conn_type][pre_index]
                    )
                    post = remove_leading_index_zero(
                        self.post_cells[conn_type][post_index]
                    )

                    if is_body_wall_muscle(post):
                        continue  # post is a BWM so ignore

                    num = self.conn_nums[conn_type][pre_index, post_index]
                    if num > 0:
                        syntype = "Send" if conn_type == HERM_CHEM else "GapJunction"
                        synclass = get_synclass(pre, syntype)

                        conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
                        # print ConnectionInfo(pre, post, num, syntype, synclass)
                        if pre not in cells:
                            cells.append(pre)
                        if post not in cells:
                            cells.append(post)

        return cells, conns

    def read_muscle_data(self):
        """
        Returns:
            neurons (:obj:`list` of :obj:`str`): List of motor neurons. Each neuron has at least one connection with a post-synaptic muscle cell.
            muscles (:obj:`list` of :obj:`str`): List of muscle cells.
            conns (:obj:`list` of :obj:`ConnectionInfo`): List of neuron-muscle connections.
        """

        neurons = []
        muscles = []
        conns = []

        """
        with open(filename, 'r') as f:
            reader = csv.DictReader(f)
            print_("Opened file: " + filename)

            for row in reader:
                pre, post, num, syntype, synclass = parse_row(row)

                if not (is_neuron(pre) or is_body_wall_muscle(pre)) or not is_body_wall_muscle(post):
                    continue

                if is_neuron(pre):
                    pre = remove_leading_index_zero(pre)
                else:
                    pre = get_old_muscle_name(pre)
                post = get_old_muscle_name(post)

                conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
                if is_neuron(pre) and pre not in neurons:
                    neurons.append(pre)
                elif is_body_wall_muscle(pre) and pre not in muscles:
                    muscles.append(pre)
                if post not in muscles:
                    muscles.append(post)"""

        return neurons, muscles, conns


def main():
    cdr = Cook2019DataReader()
    read_data = cdr.read_data
    read_muscle_data = cdr.read_muscle_data

    cells, neuron_conns = read_data(include_nonconnected_cells=True)
    neurons2muscles, muscles, muscle_conns = read_muscle_data()

    analyse_connections(cells, neuron_conns, neurons2muscles, muscles, muscle_conns)


if __name__ == "__main__":
    main()
