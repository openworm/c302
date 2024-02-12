from c302.NeuroMLUtilities import ConnectionInfo, analyse_connections
from openpyxl import load_workbook
import os
from c302 import print_

class Witvliet:
    def __init__(self, file_prefix, num_files=8):
        self.file_prefix = file_prefix
        self.num_files = num_files
        self.spreadsheet_location = os.path.dirname(os.path.abspath(__file__)) + "/data/"

    def read_data(self, include_nonconnected_cells=False, neuron_connect=False):
        all_cells = []
        all_conns = []

        for i in range(1, self.num_files + 1):
            cells = []
            conns = []
            filename = f"{self.spreadsheet_location}{self.file_prefix}_{i}.xlsx"
            wb = load_workbook(filename)
            sheet = wb.worksheets[0]
            print_("Opened the Excel file: " + filename)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                pre = str(row[0])
                post = str(row[1])
                syntype = str(row[2])
                num = int(row[3])
                synclass = 'Generic_GJ' if 'electrical' in syntype else 'Chemical_Synapse'

                conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
                if pre not in cells:
                    cells.append(pre)
                if post not in cells:
                    cells.append(post)

            if include_nonconnected_cells:
                known_nonconnected_cells = ['CANL', 'CANR', 'VC6']
                for c in known_nonconnected_cells:
                    cells.append(c)

            all_cells.append(cells)
            all_conns.append(conns)

        return all_cells, all_conns

    def read_muscle_data(self):
        all_neurons = []
        all_muscles = []
        all_conns = []

        for i in range(1, self.num_files + 1):
            neurons = []
            muscles = []
            conns = []

            filename = f"{self.spreadsheet_location}{self.file_prefix}_{i}.xlsx"
            wb = load_workbook(filename)
            sheet = wb.worksheets[0]
            print_("Opened Excel file: " + filename)

            for row in sheet.iter_rows(min_row=2, values_only=True):
                pre = str(row[0])
                post = str(row[1])
                syntype = str(row[2])
                num = int(row[3])
                synclass = 'Generic_GJ' if 'electrical' in syntype else 'Chemical_Synapse'

                conns.append(ConnectionInfo(pre, post, num, syntype, synclass))
                if pre not in neurons:
                    neurons.append(pre)
                if post not in muscles:
                    muscles.append(post)

            all_neurons.append(neurons)
            all_muscles.append(muscles)
            all_conns.append(conns)

        return all_neurons, all_muscles, all_conns

    def main(self):
        all_cells, all_neuron_conns = self.read_data(include_nonconnected_cells=True)
        all_neurons2muscles, all_muscles, all_muscle_conns = self.read_muscle_data()

        for cells, neuron_conns in zip(all_cells, all_neuron_conns):
            analyse_connections(cells, neuron_conns, [], [], [])  # Adjust the parameters accordingly

        for neurons, muscles, muscle_conns in zip(all_neurons2muscles, all_muscles, all_muscle_conns):
            analyse_connections([], [], neurons, muscles, muscle_conns)  # Adjust the parameters accordingly


if __name__ == '__main__':
    excel_reader = Witvliet(file_prefix='witvliet_2020', num_files=8)
    excel_reader.main()
